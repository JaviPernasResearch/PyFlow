from collections import deque
import openpyxl
import csv
from typing import Any, Dict, Optional, List

from ..Items.item import Item
from ..SimClock.simClock import SimClock
from .source import Source

class ScheduleSource(Source):
    """
    ScheduleSource is a class that reads scheduled events from a file or a Python dictionary
    and generates items based on the schedule.
    It supports reading from Excel (.xlsx), CSV (.csv), data (.data) files, and dictionaries.

    Arguments:
        name (str): The name of the source.
        clock (SimClock): The simulation clock.
        file_name (Optional[str]): The name of the file containing the schedule (optional if using a dictionary).
        data_dict (Optional[Dict[str, List[Any]]]): A dictionary containing schedule data (headers as keys, rows as values).
        model_item (Optional[Item]): The model item to be generated.
        sheet_name (Optional[str]): The name of the sheet to read (only for Excel files).
    """
    def __init__(self, name: str, clock: SimClock, file_name: Optional[str] = None, 
                 data_dict: Optional[Dict[str, List[Any]]] = None, 
                 model_item: Optional[Item] = None, sheet_name: Optional[str] = None):
        super().__init__(name, clock, model_item)
        
        self.file_name = file_name
        self.data_dict = data_dict
        self.file_type = file_name.split('.')[-1] if file_name else None
        self.sheet_name = sheet_name
        self.headers = None
        self.row_iterator = self._get_row_iterator()
        self.row = None
        self.current_pending_q = 0
        self.current_arrival_time = 0
        self.current_item_name = None
        self.blocked_items = deque()
        self.file = None

    def _get_row_iterator(self):
        """
        Returns an iterator over the rows of the file or dictionary based on the input type.
        Supports Excel (.xlsx), CSV (.csv), data (.data) files, and dictionaries.

        If a dictionary is provided, it uses the dictionary keys as headers and iterates over the values.
        """
        if self.data_dict:
            # If a dictionary is provided, extract headers and rows
            self.headers = list(self.data_dict.keys())
            rows = zip(*self.data_dict.values())
            return iter(rows)
        elif self.file_type == 'xlsx':
            workbook = openpyxl.load_workbook(self.file_name)
            sheet = workbook[self.sheet_name] if self.sheet_name else workbook.active
            self.headers = [cell for cell in next(sheet.iter_rows(values_only=True))]
            return sheet.iter_rows(values_only=True, min_row=2)
        elif self.file_type == 'csv':
            self.file = open(self.file_name, 'r', encoding='utf-8')
            reader = csv.reader(self.file)
            self.headers = next(reader)
            return iter(reader)
        elif self.file_type == 'data':
            self.file = open(self.file_name, 'r', encoding='utf-8')
            self.headers = self.file.readline().strip().split()
            return (line.strip().split() for line in self.file)
        else:
            raise ValueError(f"Unsupported file type: {self.file_type}")

    def start(self) -> None:
        self._schedule_next_arrival()

    def _schedule_next_arrival(self) -> None:
        try:
            row = next(self.row_iterator)
            self.row = dict(zip(self.headers, row))
            if not row or row[0] is None:
                raise StopIteration("End of file reached")

            self.current_arrival_time = float(self.row.get('Time', self.clock.get_simulation_time()))
            self.current_item_name = str(self.row.get('Name', "Default"))
            self.current_pending_q = int(self.row.get('Q', 1))

            delay = max(0, self.current_arrival_time - self.clock.get_simulation_time())
            self.clock.schedule_event(self, delay)

        except StopIteration:
            self.row = None
            if self.file:
                self.file.close()

    def unblock(self) -> bool:
        while self.blocked_items:
            the_item = self.blocked_items.popleft()
            if self.get_output().send(the_item):
                self.number_items += 1
                return True
            else:
                self.blocked_items.appendleft(the_item)
                break
        return False

    def receive(self, the_item: Item) -> bool:
        raise NotImplementedError("The Source cannot receive Items.")

    def execute(self) -> None:
        new_item = self.create_item()
        while new_item:
            if not self.get_output().send(new_item):
                self.blocked_items.append(new_item)

            self.number_items += 1
            new_item = self.create_item()
        self._schedule_next_arrival()

    def check_availability(self, the_item: Item) -> bool:
        return True

    def create_item(self) -> Optional[Item]:
        if self.current_pending_q <= 0:
            return None

        self.current_pending_q -= 1
        the_item = super().create_item(name = self.current_item_name)
        the_item.set_type(self.current_item_name)

        for header, value in list(self.row.items())[3:]:
                if header is not None and value is not None:
                    the_item.set_label_value(header, value)
                else:  
                    break

        return the_item